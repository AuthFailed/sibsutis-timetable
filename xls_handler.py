"""Модуль для работы с таблицами."""
import re

from config import db
from datetime import datetime, date
import pytz
from bs4 import BeautifulSoup as Soup
from openpyxl import load_workbook, utils as u
from openpyxl.utils import get_column_letter
import openpyxl.worksheet.worksheet as worksheet
import aiohttp

import group_list as gl

main_link = "https://sibsutis.ru/students/study/Расписание xlsx/2019-2020 - 2 полугодие"


async def get_xls_for_user(user_group) -> worksheet:
    """Выбрать/скачать файл для определённой группы."""
    if user_group in gl.mts_1_course:
        await update_schedule_files(group="МТС1к")
        wb = load_workbook(filename="./schedule_files/МТС1к.xlsx")
    elif user_group in gl.mts_2_course:
        await update_schedule_files(group="МТС2к")
        wb = load_workbook(filename="./schedule_files/МТС2к.xlsx")
    elif user_group in gl.mts_3_course:
        await update_schedule_files(group="МТС3к")
        wb = load_workbook(filename="./schedule_files/МТС3к.xlsx")
    elif user_group in gl.mts_4_course:
        await update_schedule_files(group="МТС4к")
        wb = load_workbook(filename="./schedule_files/МТС4к.xlsx")
    elif user_group in gl.mrm_1_course:
        await update_schedule_files(group="МРМ1к")
        wb = load_workbook(filename="./schedule_files/МРМ1к.xlsx")
    elif user_group in gl.mrm_2_course:
        await update_schedule_files(group="МРМ2к")
        wb = load_workbook(filename="./schedule_files/МРМ2к.xlsx")
    elif user_group in gl.mrm_3_course:
        await update_schedule_files(group="МРМ3к")
        wb = load_workbook(filename="./schedule_files/МРМ3к.xlsx")
    elif user_group in gl.mrm_4_course:
        await update_schedule_files(group="МРМ4к")
        wb = load_workbook(filename="./schedule_files/МРМ4к.xlsx")
    elif user_group in gl.ivt_1_course:
        await update_schedule_files(group="ИВТ1к")
        wb = load_workbook(filename="./schedule_files/ИВТ1к.xlsx")
    elif user_group in gl.ivt_2_course:
        await update_schedule_files(group="ИВТ2к")
        wb = load_workbook(filename="./schedule_files/ИВТ2к.xlsx")
    elif user_group in gl.ivt_3_course:
        await update_schedule_files(group="ИВТ3к")
        wb = load_workbook(filename="./schedule_files/ИВТ3к.xlsx")
    elif user_group in gl.ivt_4_course:
        await update_schedule_files(group="ИВТ4к")
        wb = load_workbook(filename="./schedule_files/ИВТ4к.xlsx")
    elif user_group in gl.gf_1_course:
        await update_schedule_files(group="ГФ1к")
        wb = load_workbook(filename="./schedule_files/ГФ1к.xlsx")
    elif user_group in gl.gf_2_course:
        await update_schedule_files(group="ГФ2к")
        wb = load_workbook(filename="./schedule_files/ГФ2к.xlsx")
    elif user_group in gl.gf_3_course:
        await update_schedule_files(group="ГФ3к")
        wb = load_workbook(filename="./schedule_files/ГФ3к.xlsx")
    elif user_group in gl.gf_4_course:
        await update_schedule_files(group="ГФ4к")
        wb = load_workbook(filename="./schedule_files/ГФ4к.xlsx")
    elif user_group in gl.aes_1_course:
        await update_schedule_files(group="АЭС1к")
        wb = load_workbook(filename="./schedule_files/АЭС1к.xlsx")
    elif user_group in gl.aes_2_course:
        await update_schedule_files(group="АЭС2к")
        wb = load_workbook(filename="./schedule_files/АЭС2к.xlsx")
    elif user_group in gl.aes_3_course:
        await update_schedule_files(group="АЭС3к")
        wb = load_workbook(filename="./schedule_files/АЭС3к.xlsx")
    elif user_group in gl.aes_4_course:
        await update_schedule_files(group="АЭС4к")
        wb = load_workbook(filename="./schedule_files/АЭС4к.xlsx")
    else:
        await update_schedule_files(group="АЭС5к")
        wb = load_workbook(filename="./schedule_files/АЭС5к.xlsx")
    wb = wb.get_sheet_by_name("TDSheet")
    return wb


async def update_schedule_files(group):
    """Обновление устаревших файлов расписания."""
    async with aiohttp.ClientSession() as session:
        async with session.get(url=main_link,
                               ssl=False) as response:
            assert response.status == 200
            soup = Soup(await response.text(), "html.parser")
            text = soup.find("a", attrs={"class": "element-title",
                                         "data-bx-title": "%s.xlsx" % group})
            parsed_time = text["data-bx-datemodify"]
            download_link = f"https://sibsutis.ru{text['data-bx-download']}"
            result = db.execute(
                query=f"SELECT update_time FROM fs WHERE file_name=\'{group}\'")
            if result != parsed_time:
                db.update_time(file_name=group, update_time=parsed_time)
                async with session.get(download_link, allow_redirects=True, ssl=False) as r:
                    open(file=f'./schedule_files/{group}.xlsx', mode='wb').write(await r.content.read())


async def get_certain_day(group, day):
    week_day = re.search(r"get_(.*?)_", day).group(1)
    if week_day == "monday":
        selected_day = get_weekday(0)
    elif week_day == "tuesday":
        selected_day = get_weekday(1)
    elif week_day == "wednesday":
        selected_day = get_weekday(2)
    elif week_day == "thursday":
        selected_day = get_weekday(3)
    elif week_day == "friday":
        selected_day = get_weekday(4)
    else:
        selected_day = get_weekday(5)

    answer_message = await get_schedule(user_group=group,
                                  day=selected_day)
    return format_message(answer_message)


async def get_today_schedule(user_group):
    """Получить сегодняшнее расписание."""
    today = get_weekday(datetime.weekday(
        datetime.now(pytz.timezone('Asia/Bangkok'))))
    try:
        if today[1] != "Воскресенье":
            answer_message = await get_schedule(user_group=user_group,
                                          day=today)
        else:
            answer_message = "*Воскресенье*\n\nСегодня выходной! Отдыхай 😊"
        return format_message(answer_message)
    except Exception as e:
        raise(e)


async def get_tomorrow_schedule(user_group) -> str:
    """Получить завтрашнее расписание."""
    if datetime.weekday(datetime.now().astimezone(pytz.timezone('Asia/Bangkok'))) + 1 == 7:
        tomorrow = get_weekday(0)
    else:
        tomorrow = get_weekday(datetime.weekday(datetime.now().astimezone(pytz.timezone('Asia/Bangkok'))) + 1)
    if tomorrow[1] != "Воскресенье":
        answer_message = await get_schedule(user_group=user_group,
                                      day=tomorrow)
    else:
        answer_message = "*Воскресенье*\n\nЗавтра выходной! Отдыхай 😊"
    return answer_message


async def get_schedule(user_group, day) -> str:
    answer_message = ''
    num_of_lessons = 0

    table = await get_xls_for_user(user_group)

    letter = await get_letter(user_group=user_group,
                        table=table)

    answer_message += "*" + day[1] + "*\n\n"
    for i in range(0, len(day[0])):
        point = letter + str(day[0][i])
        if type(table[point]).__name__ == "MergedCell":
            point = letter + str(table[point].row - 1)
        if table[point].value is not None:
            lesson_time = table["B%s" % (int(table[point].row))].value
            if type(table["B%s" % (int(table[point].row))]).__name__ == "MergedCell":
                lesson_time = table["B%s" % (int(table[point].row) - 1)].value
            info = await get_lesson_by_time(lesson_time=lesson_time, cell=table[point].value)
            answer_message += info[0]
            num_of_lessons += 1
    if len(info):
        answer_message += "\n*{0}* пар(ы). Последняя пара кончается в *{1}*.".format(num_of_lessons,
                                                                                     info[1][11:-3]
                                                                                     )
    else:
        answer_message += "\n\nВыходной! Отдыхай 😊"
    return answer_message


async def get_letter(user_group, table):
    table = table
    letter = ''
    for cellobj in table["C4":"Z4"]:
        for cell in cellobj:
            if user_group == cell.value:
                letter = get_column_letter(u.coordinate_to_tuple(cell.coordinate)[1])
                break
    return letter


async def get_lesson_by_time(lesson_time, cell):
    if lesson_time == "8:00:00 - 9:35:00":
        answer_message = "🕗 _1 пара (8:00 — 9:35)_\n" + cell + "\n"
        last_lesson_time = lesson_time
    elif lesson_time == "9:50:00 - 11:25:00":
        answer_message = "🕙 _2 пара (9:50 — 11:25)_\n" + cell + "\n"
        last_lesson_time = lesson_time
    elif lesson_time == "11:40:00 - 13:15:00":
        answer_message = "🕦 _3 пара (11:40 — 13:15)_\n" + cell + "\n"
        last_lesson_time = lesson_time
    elif lesson_time == "13:45:00 - 15:20:00":
        answer_message = "🕜 _4 пара (13:45 — 15:20)_\n" + cell + "\n"
        last_lesson_time = lesson_time
    elif lesson_time == "15:35:00 - 17:10:00":
        answer_message = "🕞 _5 пара (15:35 — 17:10)_\n" + cell + "\n"
        last_lesson_time = lesson_time
    elif lesson_time == "17:25:00 - 19:00:00":
        answer_message = "🕠 _6 пара (17:25 — 19:00)_\n" + cell + "\n"
        last_lesson_time = lesson_time
    else:
        raise Exception("Неверное время")
    arr = [answer_message, last_lesson_time]
    return arr


def format_message(text):
    """Форматируем сообщение перед отправкой."""
    text = text.replace("(Лекционные занятия)", "`(Лекция)`") \
        .replace("(Лабораторные работы)", "`(Лабораторная)`") \
        .replace("(Практические занятия)", "`(Практика)`") \
        .replace("(Занятия по ин.язу)", "`(Занятия по ин.язу)`") \
        .replace("(Занятия по физической культуре)", "`(Занятия по физической культуре)`")
    return text


def get_weekday(current_date):
    """Узнаём день недели."""
    tday = []
    if is_week_even():
        if current_date == 0:
            rows = ["5", "7", "9", "11", "13", "15"]
            day_of_week = "Понедельник"
        elif current_date == 1:
            rows = ["17", "19", "21", "23", "25", "27"]
            day_of_week = "Вторник"
        elif current_date == 2:
            rows = ["29", "31", "33", "35", "37", "39"]
            day_of_week = "Среда"
        elif current_date == 3:
            rows = ["41", "43", "45", "47", "49", "51"]
            day_of_week = "Четверг"
        elif current_date == 4:
            rows = ["53", "55", "57", "59", "61", "63"]
            day_of_week = "Пятница"
        elif current_date == 5:
            rows = ["65", "67", "69", "71", "73", "75"]
            day_of_week = "Суббота"
        else:
            rows = []
            day_of_week = "Воскресенье"
    else:
        if current_date == 0:
            rows = ["6", "8", "10", "12", "14", "16"]
            day_of_week = "Понедельник"
        elif current_date == 1:
            rows = ["18", "20", "22", "24", "26", "28"]
            day_of_week = "Вторник"
        elif current_date == 2:
            rows = ["30", "32", "34", "36", "38", "40"]
            day_of_week = "Среда"
        elif current_date == 3:
            rows = ["42", "44", "46", "48", "50", "52"]
            day_of_week = "Четверг"
        elif current_date == 4:
            rows = ["54", "56", "58", "60", "62", "64"]
            day_of_week = "Пятница"
        elif current_date == 5:
            rows = ["66", "68", "70", "72", "74", "76"]
            day_of_week = "Суббота"
        else:
            rows = []
            day_of_week = "Воскресенье"
    for obj in rows, day_of_week:
        tday.append(obj)
    print(tday)
    return tday


def get_week_from_date(date_object):
    """Получае номер учебной недели."""
    date_ordinal = date_object.toordinal()
    year = date_object.year
    week = ((date_ordinal - _week1_start_ordinal(year)) // 7) + 1
    if week >= 52 and date_ordinal > + _week1_start_ordinal(year + 1):
        year += 1
        week = 1
    return week - 34


def _week1_start_ordinal(year):
    """Устанавливаем начало года на 1 января."""
    jan1 = date(year, 1, 1)
    jan1_ordinal = jan1.toordinal()
    jan1_weekday = jan1.weekday()
    week1_start_ordinal = jan1_ordinal - ((jan1_weekday + 1) % 7)
    return week1_start_ordinal


def is_week_even():
    """Проверка недели на чётность."""
    if get_week_from_date(datetime.now()) % 2 != 0:
        return False
    else:
        return True
